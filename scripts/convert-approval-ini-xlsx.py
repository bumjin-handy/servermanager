# Optional offline converter (app loads xlsx at runtime via 불러오기).
# Output src/data/approvalIniDocs.json is gitignored — do not commit.
#
#   python scripts/convert-approval-ini-xlsx.py
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "reference" / "HANDY_HSO_Approval_ini_20201230.xlsx"
OUT = ROOT / "src" / "data" / "approvalIniDocs.json"
SKIP = {"XXXXXX", "Sheet1"}


def cell(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def trim_row(row) -> list[str]:
    vals = [cell(c) for c in row]
    while vals and vals[-1] == "":
        vals.pop()
    return vals


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    sheets_out = []

    for name in wb.sheetnames:
        if name in SKIP:
            continue
        ws = wb[name]
        if not hasattr(ws, "iter_rows"):
            continue
        raw = [trim_row(r) for r in ws.iter_rows(values_only=True)]
        raw = [r for r in raw if any(r)]

        title = name
        intro: list[str] = []
        columns: list[str] = []
        rows: list[dict] = []
        sections: list[dict] = []

        if name in (
            "globals.properties",
            "jhomscfg.xml",
            "handydef.ini",
            "기타설정및가이드",
        ):
            columns = ["옵션명", "설정값", "설명", "Default", "Version", "비고"]
            header_seen = False
            for r in raw:
                if not header_seen:
                    if r and r[0] == "옵션명":
                        header_seen = True
                        continue
                    if r and r[0]:
                        if title == name and len(r[0]) < 120:
                            title = r[0]
                        else:
                            intro.append(r[0])
                    continue
                if r and r[0] == "옵션명":
                    continue
                if r and r[0] and (
                    "Section" in r[0] or (r[0].startswith("[") and "]" in r[0])
                ):
                    sections.append({"title": r[0], "rowIndex": len(rows)})
                    continue
                padded = (r + [""] * 6)[:6]
                if not any(padded):
                    continue
                rows.append(
                    {
                        "옵션명": padded[0],
                        "설정값": padded[1],
                        "설명": padded[2],
                        "Default": padded[3],
                        "Version": padded[4],
                        "비고": padded[5],
                    }
                )

        elif name == "권한별역할":
            title = "HSO 결재관련 권한별 역할"
            columns = ["권한", "권한에 따른 역할", "메뉴", "기능", "기능권한"]
            header_seen = False
            last_auth = ""
            for r in raw:
                if not header_seen:
                    if len(r) > 1 and r[1] == "권한":
                        header_seen = True
                    continue
                auth = r[1] if len(r) > 1 else ""
                role = r[2] if len(r) > 2 else ""
                menu = r[3] if len(r) > 3 else ""
                feat = r[5] if len(r) > 5 else ""
                feat_auth = r[6] if len(r) > 6 else ""
                if auth:
                    last_auth = auth
                if not any([auth, role, menu, feat, feat_auth]):
                    continue
                rows.append(
                    {
                        "권한": auth or last_auth,
                        "권한에 따른 역할": role,
                        "메뉴": menu,
                        "기능": feat,
                        "기능권한": feat_auth,
                    }
                )

        elif name == "서식Guide":
            title = "HSO 결재 서식 셀명"
            columns = ["서식셀명", "영문셀명", "설명", "Version"]
            header_seen = False
            for r in raw:
                if not header_seen:
                    if len(r) > 1 and "서식셀명" in r[1]:
                        header_seen = True
                    continue
                padded = (r + [""] * 6)[:6]
                if not padded[1] and not padded[2]:
                    continue
                rows.append(
                    {
                        "서식셀명": padded[1],
                        "영문셀명": padded[2],
                        "설명": padded[3],
                        "Version": padded[4],
                    }
                )

        elif name == "결재알림(노티파이)":
            title = "결재 알림 (노티파이)"
            columns = ["알림종류", "알림대상", "알림내용", "내용보기", "지원버전", "기타"]
            header_seen = False
            for r in raw:
                if not header_seen:
                    if len(r) > 1 and "알림" in r[1] and "종류" in r[1]:
                        header_seen = True
                    continue
                padded = (r + [""] * 8)[:8]
                if not padded[1]:
                    continue
                rows.append(
                    {
                        "알림종류": padded[1],
                        "알림대상": padded[2],
                        "알림내용": padded[3],
                        "내용보기": padded[4],
                        "지원버전": padded[5],
                        "기타": padded[6],
                    }
                )

        elif name == "ErrorCode설명":
            title = "Handy Groupware Error Code"
            columns = ["오류번호", "심볼릭값", "의미및조치"]
            header_seen = False
            for r in raw:
                if not header_seen:
                    if r and r[0] == "오류번호":
                        header_seen = True
                    continue
                padded = (r + [""] * 3)[:3]
                if padded[0] == "" and padded[1] and padded[2] == "":
                    sections.append({"title": padded[1], "rowIndex": len(rows)})
                    continue
                if not padded[0] and not padded[1]:
                    continue
                rows.append(
                    {
                        "오류번호": padded[0],
                        "심볼릭값": padded[1],
                        "의미및조치": padded[2],
                    }
                )
        else:
            continue

        sheets_out.append(
            {
                "id": name,
                "name": name,
                "title": title,
                "intro": intro,
                "columns": columns,
                "sections": sections,
                "rows": rows,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": "HANDY_HSO_Approval_ini_20201230.xlsx",
        "sheets": sheets_out,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    for s in sheets_out:
        print(f"{s['name']}: rows={len(s['rows'])} sections={len(s['sections'])}")
    print("wrote", OUT, "(gitignored)")


if __name__ == "__main__":
    main()
